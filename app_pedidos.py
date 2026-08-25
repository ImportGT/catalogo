import os
import json
import glob
import pandas as pd
import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Diccionario completo con todas las líneas de Pandora y sus rutas exactas
CONFIGURACION_LINEAS = {
    "Anillos Pandora": {"excel": "ANILLOS.xlsx", "carpeta": "imagenes/anillos", "prefijo": "anillos", "header": 0},
    "Aretes Pandora": {"excel": "ARETES.xlsx", "carpeta": "imagenes/aretes", "prefijo": "aretes", "header": 0},
    "Collares Pandora": {"excel": "COLLARES.xlsx", "carpeta": "imagenes/collares", "prefijo": "collares", "header": 0},
    "Pulseras Pandora": {"excel": "PULSERAS.xlsx", "carpeta": "imagenes/pulseras", "prefijo": "pulseras", "header": 0},
    "Charms Beads Pandora": {"excel": "CHARMS BEADS.xlsx", "carpeta": "imagenes/charms_beads", "prefijo": "chb", "header": 0},
    "Charms Colgantes Pandora": {"excel": "CHARMS COLGANTES.xlsx", "carpeta": "imagenes/charms_colgantes", "prefijo": "chc", "header": 0, "separador": "."},
    "Charms Beads Disney Pandora": {"excel": "CHARMS BEADS DISNEY.xlsx", "carpeta": "imagenes/charms_beads_disney", "prefijo": "chbd", "header": 0},
    "Charms Colgantes Disney Pandora": {"excel": "CHARMS COLGANTES DISNEY.xlsx", "carpeta": "imagenes/charms_colgantes_disney", "prefijo": "chcd", "header": 0},
    "Charms Cadenas de Seguridad Pandora": {"excel": "CHARMS CADENAS DE SEGURIDAD.xlsx", "carpeta": "imagenes/charms_cadenasseguridad", "prefijo": "chcsd", "header": 0},
    "Charms Muranos Pandora": {"excel": "CHARMS MURANOS.xlsx", "carpeta": "imagenes/charms_muranos", "prefijo": "chm", "header": 0},
    "Charms Clips y Topes Pandora": {"excel": "CHARMS CLIPS Y TOPES.xlsx", "carpeta": "imagenes/charms_clipsytopes", "prefijo": "chct", "header": 0},
    "Charms Locket Pandora": {"excel": "CHARMS LOCKET.xlsx", "carpeta": "imagenes/charms_lockets", "prefijo": "chl", "header": 0},
    "Charms Reflexion Pandora": {"excel": "CHARMS REFLEXION.xlsx", "carpeta": "imagenes/charms_reflexion", "prefijo": "chr", "header": 0},
    "Charms ME Pandora": {"excel": "CHARMS ME.xlsx", "carpeta": "imagenes/charms_me", "prefijo": "chme", "header": 0},
    "Charms Accesorios ME Pandora": {"excel": "CHARMS ACCESORIOS ME.xlsx", "carpeta": "imagenes/charms_accesoriosme", "prefijo": "chamE", "header": 0}
}

CARPETA_GUARDADOS = "pedidos_guardados"

class AppGestorPedidos:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Pedidos y Cotizaciones - Joyería")
        self.root.geometry("1280x850")
        self.root.config(bg="#f4f6f9")

        self.items_pedido = []
        self.imagenes_referencias = []
        self.nombre_archivo_actual = None

        if not os.path.exists(CARPETA_GUARDADOS):
            os.makedirs(CARPETA_GUARDADOS)

        # --- CONTENEDOR PRINCIPAL EN DOS COLUMNAS ---
        main_container = tk.Frame(root, bg="#f4f6f9")
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # COLUMNA IZQUIERDA: GESTIÓN DE PEDIDOS Y COTIZACIONES
        left_frame = tk.Frame(main_container, bg="#f4f6f9")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # COLUMNA DERECHA: PANEL FIJO DE INVENTARIO Y STOCK
        right_frame = tk.LabelFrame(main_container, text=" Inventario y Stock en Tienda ", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#1F4E79", padx=5, pady=5)
        right_frame.pack(side="right", fill="y", padx=(5, 0))

        # Configuración de la tabla visual (Treeview) para el inventario
        columns_inv = ("linea", "stock")
        self.tree_inventario = ttk.Treeview(right_frame, columns=columns_inv, show="headings", height=30)
        self.tree_inventario.heading("linea", text="Línea / Categoría")
        self.tree_inventario.heading("stock", text="Stock Total")
        self.tree_inventario.column("linea", width=190, anchor="w")
        self.tree_inventario.column("stock", width=80, anchor="center")
        
        scrollbar_inv = ttk.Scrollbar(right_frame, orient="vertical", command=self.tree_inventario.yview)
        self.tree_inventario.configure(yscrollcommand=scrollbar_inv.set)

        self.tree_inventario.pack(side="left", fill="y", expand=True)
        scrollbar_inv.pack(side="right", fill="y")
        self.cargar_datos_inventario_pantalla()

        # --- SECCIÓN SUPERIOR IZQUIERDA: RECUPERAR / GUARDAR PEDIDOS ---
        frame_recuperar = tk.LabelFrame(left_frame, text=" Administrar Cotizaciones Guardadas ", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#1F4E79", padx=10, pady=6)
        frame_recuperar.pack(fill="x", padx=5, pady=5)

        tk.Label(frame_recuperar, text="Historial:", bg="#f4f6f9", font=("Helvetica", 9)).grid(row=0, column=0, sticky="w", padx=2)
        self.combo_guardados = ttk.Combobox(frame_recuperar, width=28, state="readonly", font=("Helvetica", 9))
        self.combo_guardados.grid(row=0, column=1, padx=2, sticky="w")
        self.actualizar_combo_guardados()

        btn_cargar = tk.Button(frame_recuperar, text="Cargar", bg="#1F4E79", fg="white", font=("Helvetica", 9, "bold"), command=self.cargar_pedido_guardado)
        btn_cargar.grid(row=0, column=2, padx=4, sticky="w")

        btn_nuevo = tk.Button(frame_recuperar, text="Nuevo", bg="#008000", fg="white", font=("Helvetica", 9, "bold"), command=self.nuevo_pedido)
        btn_nuevo.grid(row=0, column=3, padx=2, sticky="w")

        # --- SECCIÓN: DATOS DEL CLIENTE Y PAGOS ---
        frame_top = tk.LabelFrame(left_frame, text=" Datos del Cliente, Márgenes y Anticipo ", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#1F4E79", padx=10, pady=8)
        frame_top.pack(fill="x", padx=5, pady=5)

        tk.Label(frame_top, text="Cliente:", bg="#f4f6f9", font=("Helvetica", 9, "bold")).grid(row=0, column=0, sticky="w", padx=2)
        self.entry_cliente = tk.Entry(frame_top, width=18, font=("Helvetica", 10))
        self.entry_cliente.grid(row=0, column=1, padx=2, sticky="w")
        self.entry_cliente.bind("<KeyRelease>", lambda e: self.guardar_estado_actual())

        tk.Label(frame_top, text="Margen(%):", bg="#f4f6f9", font=("Helvetica", 9, "bold")).grid(row=0, column=2, sticky="w", padx=5)
        self.entry_margen = tk.Entry(frame_top, width=5, font=("Helvetica", 10))
        self.entry_margen.insert(0, "0")
        self.entry_margen.grid(row=0, column=3, padx=2, sticky="w")
        self.entry_margen.bind("<KeyRelease>", lambda e: self.actualizar_totales_pantalla())

        tk.Label(frame_top, text="Anticipo(Q):", bg="#f4f6f9", font=("Helvetica", 9, "bold")).grid(row=0, column=4, sticky="w", padx=5)
        self.entry_anticipo = tk.Entry(frame_top, width=8, font=("Helvetica", 10))
        self.entry_anticipo.insert(0, "0.00")
        self.entry_anticipo.grid(row=0, column=5, padx=2, sticky="w")
        self.entry_anticipo.bind("<KeyRelease>", lambda e: self.actualizar_totales_pantalla())

        # --- SECCIÓN MEDIA: ENTRADA DE PRODUCTOS ---
        frame_mid = tk.LabelFrame(left_frame, text=" Agregar Producto ", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#1F4E79", padx=10, pady=8)
        frame_mid.pack(fill="x", padx=5, pady=5)

        tk.Label(frame_mid, text="Categoría / Línea:", bg="#f4f6f9", font=("Helvetica", 9)).grid(row=0, column=0, sticky="w", padx=2)
        self.combo_linea = ttk.Combobox(frame_mid, values=list(CONFIGURACION_LINEAS.keys()), width=28, state="readonly", font=("Helvetica", 9))
        self.combo_linea.grid(row=0, column=1, padx=2, sticky="w")
        if CONFIGURACION_LINEAS:
            self.combo_linea.current(0)

        tk.Label(frame_mid, text="Código:", bg="#f4f6f9", font=("Helvetica", 9)).grid(row=0, column=2, sticky="w", padx=5)
        self.entry_codigo = tk.Entry(frame_mid, width=8, font=("Helvetica", 10))
        self.entry_codigo.grid(row=0, column=3, padx=2, sticky="w")

        tk.Label(frame_mid, text="Talla:", bg="#f4f6f9", font=("Helvetica", 9)).grid(row=1, column=0, sticky="w", padx=2, pady=6)
        self.entry_talla = tk.Entry(frame_mid, width=12, font=("Helvetica", 10))
        self.entry_talla.insert(0, "única")
        self.entry_talla.grid(row=1, column=1, padx=2, sticky="w", pady=6)

        tk.Label(frame_mid, text="Cant:", bg="#f4f6f9", font=("Helvetica", 9)).grid(row=1, column=2, sticky="w", padx=5, pady=6)
        self.entry_qty = tk.Entry(frame_mid, width=6, font=("Helvetica", 10))
        self.entry_qty.insert(0, "1")
        self.entry_qty.grid(row=1, column=3, padx=2, sticky="w", pady=6)

        btn_agregar = tk.Button(frame_mid, text="Agregar", bg="#1F4E79", fg="white", font=("Helvetica", 9, "bold"), padx=10, command=self.agregar_producto)
        btn_agregar.grid(row=0, column=4, rowspan=2, padx=10, sticky="nsew")

        # --- SECCIÓN LISTA ENLISTADA CON MINIATURAS ---
        frame_lista_container = tk.LabelFrame(left_frame, text=" Productos en el Pedido Actual ", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#1F4E79", padx=5, pady=5)
        frame_lista_container.pack(fill="both", expand=True, padx=5, pady=5)

        self.canvas = tk.Canvas(frame_lista_container, bg="white", highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(frame_lista_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="white")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # --- SECCIÓN INFERIOR: TOTALES Y BOTONES ---
        frame_bottom_panel = tk.Frame(left_frame, bg="#f4f6f9", padx=2, pady=5)
        frame_bottom_panel.pack(fill="x", padx=5, pady=5)

        self.lbl_total_general = tk.Label(frame_bottom_panel, text="Total: Q0.00", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#008000")
        self.lbl_total_general.pack(side="left", padx=2)

        self.lbl_saldo_pendiente = tk.Label(frame_bottom_panel, text="Saldo: Q0.00", font=("Helvetica", 10, "bold"), bg="#f4f6f9", fg="#C00000")
        self.lbl_saldo_pendiente.pack(side="left", padx=5)

        btn_generar_orden = tk.Button(frame_bottom_panel, text="OC (Excel)", bg="#1F4E79", fg="white", font=("Helvetica", 9, "bold"), padx=6, pady=4, command=self.generar_excel_proveedor)
        btn_generar_orden.pack(side="right", padx=2)

        btn_generar_pdf = tk.Button(frame_bottom_panel, text="Cotización (PDF)", bg="#008000", fg="white", font=("Helvetica", 9, "bold"), padx=6, pady=4, command=self.generar_pdf_cliente)
        btn_generar_pdf.pack(side="right", padx=2)

        btn_inventario = tk.Button(frame_bottom_panel, text="Reporte Inv.", bg="#595959", fg="white", font=("Helvetica", 9, "bold"), padx=6, pady=4, command=self.generar_reporte_inventario)
        btn_inventario.pack(side="right", padx=2)

    def cargar_datos_inventario_pantalla(self):
        for row in self.tree_inventario.get_children():
            self.tree_inventario.delete(row)

        for nombre_linea, cfg in CONFIGURACION_LINEAS.items():
            total_stock = 0.0
            if os.path.exists(cfg["excel"]):
                try:
                    df = pd.read_excel(cfg["excel"], header=cfg["header"])
                    df.columns = [str(c).strip().upper() for c in df.columns]
                    
                    # Buscar todas las columnas que contengan la palabra "STOCK" (ej. STOCK 7, STOCK 8, etc.)
                    columnas_stock = [c for c in df.columns if "STOCK" in c]
                    
                    if columnas_stock:
                        for col in columnas_stock:
                            for val in df[col]:
                                if not pd.isna(val):
                                    try:
                                        total_stock += float(val)
                                    except (ValueError, TypeError):
                                        pass
                except Exception:
                    pass
            
            self.tree_inventario.insert("", "end", values=(nombre_linea, f"{total_stock:.0f}"))

    def actualizar_combo_guardados(self):
        if os.path.exists(CARPETA_GUARDADOS):
            archivos = [f.replace(".json", "") for f in os.listdir(CARPETA_GUARDADOS) if f.endswith(".json")]
            archivos.sort(reverse=True)
            self.combo_guardados['values'] = archivos
            if archivos and not self.combo_guardados.get():
                self.combo_guardados.current(0)

    def nuevo_pedido(self):
        self.items_pedido.clear()
        self.refrescar_lista_visual()
        self.entry_cliente.delete(0, tk.END)
        self.entry_anticipo.delete(0, tk.END)
        self.entry_anticipo.insert(0, "0.00")
        self.nombre_archivo_actual = None
        messagebox.showinfo("Nuevo", "Listo para crear una nueva cotización.")

    def guardar_estado_actual(self):
        cliente = self.entry_cliente.get().strip()
        if not cliente or not self.items_pedido:
            return

        fecha_hoy = pd.Timestamp.now().strftime('%d%m%y')

        if not self.nombre_archivo_actual or not self.nombre_archivo_actual.startswith(cliente):
            patron = os.path.join(CARPETA_GUARDADOS, f"{cliente}-*-*.json")
            existentes = glob.glob(patron)
            
            patron_hoy = os.path.join(CARPETA_GUARDADOS, f"{cliente}-*-{fecha_hoy}.json")
            existentes_hoy = glob.glob(patron_hoy)

            if existentes_hoy:
                self.nombre_archivo_actual = os.path.basename(existentes_hoy[0]).replace(".json", "")
            else:
                max_num = 0
                for f in existentes:
                    nombre_base = os.path.basename(f).replace(".json", "")
                    partes = nombre_base.split("-")
                    if len(partes) >= 2 and partes[1].isdigit():
                        num = int(partes[1])
                        if num > max_num:
                            max_num = num
                nuevo_num = max_num + 1
                self.nombre_archivo_actual = f"{cliente}-{nuevo_num}-{fecha_hoy}"

        ruta_completa = os.path.join(CARPETA_GUARDADOS, f"{self.nombre_archivo_actual}.json")
        
        datos_estado = {
            "cliente": cliente,
            "margen": self.entry_margen.get(),
            "anticipo": self.entry_anticipo.get(),
            "items": self.items_pedido
        }

        try:
            with open(ruta_completa, 'w', encoding='utf-8') as f:
                json.dump(datos_estado, f, ensure_ascii=False, indent=4)
            self.actualizar_combo_guardados()
        except Exception:
            pass

    def cargar_pedido_guardado(self):
        seleccion = self.combo_guardados.get()
        if not seleccion:
            messagebox.showwarning("Atención", "Selecciona una cotización guardada de la lista.")
            return

        archivo_guardado = os.path.join(CARPETA_GUARDADOS, f"{seleccion}.json")
        if not os.path.exists(archivo_guardado):
            messagebox.showerror("Error", "El archivo de respaldo no existe.")
            return

        try:
            with open(archivo_guardado, 'r', encoding='utf-8') as f:
                datos = json.load(f)

            self.nombre_archivo_actual = seleccion
            self.entry_cliente.delete(0, tk.END)
            self.entry_cliente.insert(0, datos.get("cliente", ""))

            self.entry_margen.delete(0, tk.END)
            self.entry_margen.insert(0, datos.get("margen", "0"))

            self.entry_anticipo.delete(0, tk.END)
            self.entry_anticipo.insert(0, datos.get("anticipo", "0.00"))

            self.items_pedido = datos.get("items", [])
            self.refrescar_lista_visual()
            messagebox.showinfo("Éxito", f"Cotización de '{datos.get('cliente')}' cargada correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el archivo: {e}")

    def buscar_en_excel(self, nombre_linea, prod_id):
        cfg = CONFIGURACION_LINEAS[nombre_linea]
        if not os.path.exists(cfg["excel"]):
            return 0.0, "", 0.0

        try:
            df = pd.read_excel(cfg["excel"], header=cfg["header"])
            df.columns = [str(c).strip().upper() for c in df.columns]
        except Exception:
            return 0.0, "", 0.0

        for _, row in df.iterrows():
            raw_id = None
            for col_id in ['NUMERO', 'NUM.', 'CODIGO', 'CÓDIGO']:
                matches = [c for c in df.columns if c == col_id]
                if matches and matches[0] in row and not pd.isna(row[matches[0]]):
                    raw_id = row[matches[0]]
                    break
            if raw_id is None:
                raw_id = row.iloc[0]

            try:
                current_id = int(float(raw_id))
            except (ValueError, TypeError):
                continue

            if current_id == int(prod_id):
                precio_base = 0.0
                for col_p in ['PRECIO', 'VENTA MAYORISTA', 'PRECIO MAYORISTA', 'PRECIO DOBLE']:
                    matches_p = [c for c in df.columns if c == col_p]
                    if matches_p and matches_p[0] in row and not pd.isna(row[matches_p[0]]):
                        try:
                            precio_base = float(row[matches_p[0]])
                            break
                        except (ValueError, TypeError):
                            pass

                # Sumar el stock de todas las columnas que contengan "STOCK" para este producto específico
                stock_disponible = 0.0
                columnas_stock = [c for c in df.columns if "STOCK" in c]
                for col in columnas_stock:
                    if not pd.isna(row[col]):
                        try:
                            stock_disponible += float(row[col])
                        except (ValueError, TypeError):
                            pass

                carpeta = cfg["carpeta"]
                prefijo = cfg["prefijo"]
                separador = cfg.get("separador", "_")
                
                extensiones = ['.jpg', '.jpeg', '.avif', '.webp', '.png']
                patrones_sufijos = ['', '.1', '.0', '_1', '_0']
                
                imagen_path = ""
                for suf in patrones_sufijos:
                    for ext in extensiones:
                        prueba_ruta = f"{carpeta}/{prefijo}{separador}{prod_id}{suf}{ext}"
                        if os.path.exists(prueba_ruta):
                            imagen_path = prueba_ruta
                            break
                    if imagen_path:
                        break

                return precio_base, imagen_path, stock_disponible

        return 0.0, "", 0.0

    def agregar_producto(self):
        linea = self.combo_linea.get()
        codigo_str = self.entry_codigo.get().strip()
        talla = self.entry_talla.get().strip() or "única"
        qty_str = self.entry_qty.get().strip()

        if not codigo_str:
            messagebox.showerror("Error", "Debes ingresar un código de producto.")
            return

        try:
            prod_id = int(codigo_str)
            qty = int(qty_str)
        except ValueError:
            messagebox.showerror("Error", "El código y la cantidad deben ser números enteros.")
            return

        precio_base, imagen_encontrada, stock_disponible = self.buscar_en_excel(linea, prod_id)

        if precio_base == 0.0:
            if not messagebox.askyesno("Aviso", f"No se encontró precio automático para el código {prod_id}. ¿Deseas ingresarlo manualmente?"):
                return
            precio_base = float(messagebox.askstring("Precio Manual", "Ingresa el precio base en Q:") or 0)

        if stock_disponible > 0:
            messagebox.showwarning("¡Alerta de Stock!", f"⚠️ El producto con código {prod_id} ya cuenta con {stock_disponible:.0f} unidades en total (en sus distintas tallas/columnas de stock).")

        item = {
            "id": prod_id,
            "categoria": linea,
            "talla": talla,
            "qty": qty,
            "precio_base": precio_base,
            "imagen": imagen_encontrada
        }

        self.items_pedido.append(item)
        self.refrescar_lista_visual()
        self.guardar_estado_actual()
        
        self.entry_codigo.delete(0, tk.END)
        self.entry_codigo.focus()

    def refrescar_lista_visual(self):
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        self.imagenes_referencias.clear()

        for index, item in enumerate(self.items_pedido):
            card = tk.Frame(self.scrollable_frame, bg="#ffffff", bd=1, relief="solid", padx=4, pady=4)
            card.pack(fill="x", padx=4, pady=3)

            lbl_img = tk.Label(card, bg="#f0f0f0", width=40, height=40)
            if item["imagen"] and os.path.exists(item["imagen"]):
                try:
                    img_pil = Image.open(item["imagen"])
                    img_pil = img_pil.resize((35, 35), Image.Resampling.LANCZOS)
                    img_tk = ImageTk.PhotoImage(img_pil)
                    self.imagenes_referencias.append(img_tk)
                    lbl_img.config(image=img_tk, text="")
                except Exception:
                    lbl_img.config(text="[Sin img]")
            else:
                lbl_img.config(text="[Sin img]", font=("Helvetica", 6))
            lbl_img.pack(side="left", padx=4)

            info_texto = f"Cod: {item['id']} | {item['categoria'][:14]} | T: {item['talla']} | Cant: {item['qty']} | Q{item['precio_base']:.2f}"
            lbl_info = tk.Label(card, text=info_texto, bg="#ffffff", font=("Helvetica", 8, "bold"), fg="#333333")
            lbl_info.pack(side="left", padx=6)

            btn_eliminar_item = tk.Button(card, text="X", bg="#C00000", fg="white", font=("Helvetica", 7, "bold"), padx=4, command=lambda idx=index: self.eliminar_producto(idx))
            btn_eliminar_item.pack(side="right", padx=4)

        self.actualizar_totales_pantalla()

    def eliminar_producto(self, index):
        del self.items_pedido[index]
        self.refrescar_lista_visual()
        self.guardar_estado_actual()

    def actualizar_totales_pantalla(self):
        try:
            margen = float(self.entry_margen.get().strip() or 0)
        except ValueError:
            margen = 0.0

        try:
            anticipo = float(self.entry_anticipo.get().strip() or 0)
        except ValueError:
            anticipo = 0.0

        total_general = 0.0
        for item in self.items_pedido:
            precio_final = item['precio_base'] * (1 + (margen / 100))
            total_general += precio_final * item['qty']

        saldo_pendiente = total_general - anticipo

        self.lbl_total_general.config(text=f"Total: Q{total_general:.2f}")
        self.lbl_saldo_pendiente.config(text=f"Saldo: Q{saldo_pendiente:.2f}")
        self.guardar_estado_actual()

    def generar_reporte_inventario(self):
        try:
            self.cargar_datos_inventario_pantalla()
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario y Reposición"

            font_titulo = Font(name="Calibri", size=13, bold=True, color="1F4E79")
            font_sub = Font(name="Calibri", size=10, bold=True, color="555555")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0')
            )

            ws.cell(row=1, column=1, value="REPORTE GENERAL DE INVENTARIO Y REPOSICIÓN").font = font_titulo
            ws.row_dimensions[1].height = 24
            ws.cell(row=2, column=1, value=f"Fecha de corte: {pd.Timestamp.now().strftime('%Y-%m-%d')}").font = font_sub
            ws.row_dimensions[2].height = 18

            headers = ["Línea / Categoría", "Total Referencias", "Total Piezas en Stock", "Sugerencia de Pedido"]
            header_row = 4
            
            for col_num, h_text in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_num, value=h_text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[header_row].height = 25

            start_row = 5
            row_idx = start_row

            for nombre_linea, cfg in CONFIGURACION_LINEAS.items():
                total_refs = 0
                total_stock = 0.0

                if os.path.exists(cfg["excel"]):
                    try:
                        df = pd.read_excel(cfg["excel"], header=cfg["header"])
                        df.columns = [str(c).strip().upper() for c in df.columns]
                        total_refs = len(df)

                        columnas_stock = [c for c in df.columns if "STOCK" in c]
                        for col in columnas_stock:
                            for val in df[col]:
                                if not pd.isna(val):
                                    try:
                                        total_stock += float(val)
                                    except (ValueError, TypeError):
                                        pass
                    except Exception:
                        pass

                ws.row_dimensions[row_idx].height = 20
                c1 = ws.cell(row=row_idx, column=1, value=nombre_linea)
                c2 = ws.cell(row=row_idx, column=2, value=total_refs)
                c3 = ws.cell(row=row_idx, column=3, value=total_stock)
                c4 = ws.cell(row=row_idx, column=4, value=f'=IF(C{row_idx}<5, "REVISAR / PEDIR", "OK")')

                c1.alignment = align_left
                c2.alignment = align_center
                c3.alignment = align_center
                c4.alignment = align_center

                for c in [c1, c2, c3, c4]:
                    c.border = thin_border
                    c.font = Font(name="Calibri", size=10)

                row_idx += 1

            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 30

            carpeta_reportes = "REPORTES INVENTARIO"
            if not os.path.exists(carpeta_reportes):
                os.makedirs(carpeta_reportes)

            archivo_reporte = os.path.join(carpeta_reportes, f"Inventario_Stock_{pd.Timestamp.now().strftime('%d%m%y')}.xlsx")
            wb.save(archivo_reporte)
            messagebox.showinfo("¡Inventario Actualizado!", f"Reporte generado con éxito:\n\n{archivo_reporte}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error: {e}")

    def generar_pdf_cliente(self):
        cliente = self.entry_cliente.get().strip()
        if not cliente:
            messagebox.showerror("Error", "Por favor ingresa el nombre del cliente.")
            return

        if not self.items_pedido:
            messagebox.showerror("Error", "La lista de productos está vacía.")
            return

        try:
            margen = float(self.entry_margen.get().strip() or 0)
            anticipo = float(self.entry_anticipo.get().strip() or 0)
            self.guardar_estado_actual()

            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            
            pdf.set_font('helvetica', 'B', 16)
            pdf.set_text_color(31, 78, 121)
            pdf.cell(0, 10, "COTIZACIÓN DE PRODUCTOS", new_x="LMARGIN", new_y="NEXT", align='C')
            
            pdf.set_font('helvetica', '', 10)
            pdf.set_text_color(90, 100, 110)
            pdf.cell(0, 6, f"Cliente: {cliente}", new_x="LMARGIN", new_y="NEXT", align='L')
            pdf.cell(0, 6, f"Fecha: {pd.Timestamp.now().strftime('%Y-%m-%d')}", new_x="LMARGIN", new_y="NEXT", align='L')
            pdf.ln(4)
            
            pdf.set_fill_color(31, 78, 121)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('helvetica', 'B', 10)
            
            pdf.cell(18, 10, "Img", 1, 0, 'C', True)
            pdf.cell(18, 10, "Cód.", 1, 0, 'C', True)
            pdf.cell(64, 10, "Categoría", 1, 0, 'C', True)
            pdf.cell(30, 10, "Detalle/Talla", 1, 0, 'C', True)
            pdf.cell(18, 10, "Cant.", 1, 0, 'C', True)
            pdf.cell(21, 10, "P. Unit.", 1, 0, 'C', True)
            pdf.cell(21, 10, "Subtotal", 1, 1, 'C', True)

            pdf.set_font('helvetica', '', 9)
            pdf.set_text_color(0, 0, 0)
            total_general = 0.0

            for item in self.items_pedido:
                current_y = pdf.get_y()
                if current_y > 260:
                    pdf.add_page()
                    current_y = pdf.get_y()

                pdf.cell(18, 14, "", 1, 0, 'C')
                pdf.cell(18, 14, str(item['id']), 1, 0, 'C')
                pdf.cell(64, 14, item['categoria'], 1, 0, 'L')
                pdf.cell(30, 14, str(item['talla']), 1, 0, 'C')
                pdf.cell(18, 14, str(item['qty']), 1, 0, 'C')
                
                precio_final = item['precio_base'] * (1 + (margen / 100))
                subtotal = precio_final * item['qty']
                total_general += subtotal

                pdf.cell(21, 14, f"Q{precio_final:.2f}", 1, 0, 'R')
                pdf.cell(21, 14, f"Q{subtotal:.2f}", 1, 1, 'R')

                if item['imagen'] and os.path.exists(item['imagen']):
                    try:
                        pdf.image(item['imagen'], x=pdf.get_x() + 2, y=current_y + 2, w=14, h=10)
                    except Exception:
                        pass

            saldo_pendiente = total_general - anticipo
            pdf.ln(4)
            pdf.set_font('helvetica', 'B', 11)
            pdf.set_text_color(0, 128, 0)
            pdf.cell(0, 6, f"TOTAL GENERAL: Q{total_general:.2f}", align='R', new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(31, 78, 121)
            pdf.cell(0, 6, f"ANTICIPO: Q{anticipo:.2f}", align='R', new_x="LMARGIN", new_y="NEXT")
            
            if saldo_pendiente > 0:
                pdf.set_text_color(192, 0, 0)
            else:
                pdf.set_text_color(0, 128, 0)
            pdf.cell(0, 6, f"SALDO PENDIENTE: Q{saldo_pendiente:.2f}", align='R', new_x="LMARGIN", new_y="NEXT")

            carpeta_pdf = "COTIZACIONES CLIENTES"
            if not os.path.exists(carpeta_pdf):
                os.makedirs(carpeta_pdf)

            base_nombre = self.nombre_archivo_actual if self.nombre_archivo_actual else f"{cliente}-1-{pd.Timestamp.now().strftime('%d%m%y')}"
            archivo_pdf = os.path.join(carpeta_pdf, f"{base_nombre}.pdf")
            pdf.output(archivo_pdf)
            messagebox.showinfo("¡Éxito!", f"Cotización generada:\n\n{archivo_pdf}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar el PDF: {e}")

    def generar_excel_proveedor(self):
        cliente = self.entry_cliente.get().strip()
        if not cliente:
            messagebox.showerror("Error", "Por favor ingresa el nombre del cliente.")
            return

        if not self.items_pedido:
            messagebox.showerror("Error", "La lista de productos está vacía.")
            return

        try:
            self.guardar_estado_actual()
            wb = Workbook()
            ws = wb.active
            ws.title = "Orden de Compra"

            font_titulo = Font(name="Calibri", size=13, bold=True, color="1F4E79")
            font_sub = Font(name="Calibri", size=11, bold=True, color="333333")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0')
            )
            bold_border = Border(
                left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                top=Side(style='thin', color='B0B0B0'), bottom=Side(style='double', color='000000')
            )

            ws.cell(row=1, column=1, value="ORDEN DE COMPRA / PROVEEDOR").font = font_titulo
            ws.row_dimensions[1].height = 24
            ws.cell(row=2, column=1, value=f"Cliente: {cliente}").font = font_sub
            ws.row_dimensions[2].height = 20
            ws.cell(row=3, column=1, value=f"Fecha: {pd.Timestamp.now().strftime('%Y-%m-%d')}").font = font_sub
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 10

            headers = ["Imagen", "Código", "Categoría", "Talla/Medida", "Cantidad", "Precio Ind. USD", "Precio Total USD"]
            header_row = 5
            
            for col_num, h_text in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_num, value=h_text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[header_row].height = 25

            start_row = 6
            for idx, item in enumerate(self.items_pedido):
                row_idx = start_row + idx
                ws.row_dimensions[row_idx].height = 90
                
                c1 = ws.cell(row=row_idx, column=1, value="")
                c2 = ws.cell(row=row_idx, column=2, value=item['id'])
                c3 = ws.cell(row=row_idx, column=3, value=item['categoria'])
                c4 = ws.cell(row=row_idx, column=4, value=item['talla'])
                c5 = ws.cell(row=row_idx, column=5, value=item['qty'])
                c6 = ws.cell(row=row_idx, column=6, value="")
                c7 = ws.cell(row=row_idx, column=7, value=f"=E{row_idx}*F{row_idx}")

                c1.alignment = align_center
                c2.alignment = align_center
                c3.alignment = align_left
                c4.alignment = align_center
                c5.alignment = align_center
                c6.alignment = align_center
                c7.alignment = align_center

                for c in [c1, c2, c3, c4, c5, c6, c7]:
                    c.border = thin_border
                    c.font = Font(name="Calibri", size=10)

                if item['imagen'] and os.path.exists(item['imagen']):
                    try:
                        img_xl = OpenpyxlImage(item['imagen'])
                        img_xl.width = 82
                        img_xl.height = 82
                        img_xl.left = 3
                        img_xl.top = 2
                        ws.add_image(img_xl, f"A{row_idx}")
                    except Exception:
                        pass

            end_row = start_row + len(self.items_pedido) - 1
            totales_row = end_row + 1
            ws.row_dimensions[totales_row].height = 22
            
            for col_num in range(1, 8):
                cell = ws.cell(row=totales_row, column=col_num)
                cell.border = bold_border
                cell.font = Font(name="Calibri", size=10, bold=True)

            ws.cell(row=totales_row, column=3, value="TOTALES:").alignment = Alignment(horizontal="right", vertical="center")
            ws.cell(row=totales_row, column=5, value=f"=SUM(E{start_row}:E{end_row})").alignment = align_center
            ws.cell(row=totales_row, column=7, value=f"=SUM(G{start_row}:G{end_row})").alignment = align_center

            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 28
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 16

            carpeta_excel = "OC"
            if not os.path.exists(carpeta_excel):
                os.makedirs(carpeta_excel)

            base_nombre = self.nombre_archivo_actual if self.nombre_archivo_actual else f"{cliente}-1-{pd.Timestamp.now().strftime('%d%m%y')}"
            nombre_oc = f"OC-{base_nombre}.xlsx"
            archivo_excel = os.path.join(carpeta_excel, nombre_oc)
            wb.save(archivo_excel)
            messagebox.showinfo("¡Éxito!", f"Orden de compra generada:\n\n{archivo_excel}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar la orden: {e}")

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = AppGestorPedidos(root)
        root.mainloop()
    except Exception as e:
        root_err = tk.Tk()
        root_err.withdraw()
        messagebox.showerror("Error Crítico de Inicio", f"El programa no pudo arrancar debido a:\n{str(e)}")